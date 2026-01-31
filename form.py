"""
Output formatters for Copy Block Generator.
Exports results to CSV, JSON, and Excel with flattened structure for easy validation.
"""

import csv
import json
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class OutputFormatter:
    """Format and save workflow outputs to various file formats."""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def save_all(self, result: Dict[str, Any], base_name: str = None) -> Dict[str, Path]:
        """Save result to all formats: JSON, CSV, and Excel."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        query = result.get('query', 'output')
        safe_query = "".join(c if c.isalnum() or c in ' -_' else '_' for c in query)[:30]
        base_name = base_name or f"{safe_query}_{timestamp}"
        
        paths = {}
        
        # JSON - full output
        paths['json'] = self.save_json(result, base_name)
        
        # CSV - flattened
        paths['csv'] = self.save_csv(result, base_name)
        
        # Excel - validation friendly
        if EXCEL_AVAILABLE:
            paths['excel'] = self.save_excel(result, base_name)
        
        return paths
    
    def save_json(self, result: Dict[str, Any], base_name: str) -> Path:
        """Save complete result as JSON."""
        path = self.output_dir / f"{base_name}.json"
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, default=str)
        return path
    
    def save_excel(self, result: Dict[str, Any], base_name: str) -> Path:
        """Save result as Excel with flattened single-row structure."""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl required for Excel export. Install: pip install openpyxl")
        
        path = self.output_dir / f"{base_name}.xlsx"
        
        # Flatten
        row_data = self._flatten_result_to_single_row(result)
        
        # Write single row using common helper
        self._write_excel_file(path, [row_data], title="Results")
        return path

    def save_batch_report(self, results: List[Dict[str, Any]], base_name: str = "batch_report") -> Dict[str, Path]:
        """Save a batch of results into single consolidated CSV and Excel files."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_name = f"{base_name}_{timestamp}"
        paths = {}

        flattened_rows = [self._flatten_result_to_single_row(r) for r in results]
        
        if not flattened_rows:
            return paths

        # Save CSV
        csv_path = self.output_dir / f"{final_name}.csv"
        headers = list(flattened_rows[0].keys())
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(flattened_rows)
        paths['csv'] = csv_path

        # Save Excel
        if EXCEL_AVAILABLE:
            xlsx_path = self.output_dir / f"{final_name}.xlsx"
            self._write_excel_file(xlsx_path, flattened_rows, title="Batch Results")
            paths['excel'] = xlsx_path
            
        return paths

    def _write_excel_file(self, path: Path, rows: List[Dict[str, Any]], title: str = "Results"):
        """Helper to write a list of flattened dictionaries to an Excel file with styling."""
        wb = Workbook()
        ws = wb.active
        ws.title = title
        
        if not rows:
            wb.save(path)
            return

        headers = list(rows[0].keys())
        styles = self._get_excel_styles()
        
        # Write Header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.border = styles['border']
        
        # Write Data
        for r_idx, row_data in enumerate(rows, 2):
            for c_idx, header in enumerate(headers, 1):
                val = row_data.get(header, '')
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = styles['alignment']
                cell.border = styles['border']
                
                # Highlight issues
                if header == 'Overall Verdict' and val != 'PASS':
                    cell.fill = styles['warning_fill']
        
        # Adjust widths
        self._adjust_column_widths(ws, headers)
        
        wb.save(path)

    def _get_excel_styles(self) -> Dict[str, Any]:
        """Return common Excel styles."""
        return {
            'header_font': Font(bold=True, color="FFFFFF"),
            'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'warning_fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin')),
            'alignment': Alignment(wrap_text=True, vertical="top")
        }

    def _adjust_column_widths(self, ws, headers):
        """Set nice column widths based on header name."""
        for col, header in enumerate(headers, 1):
            width = 30
            if 'Copy' in header: width = 80
            if 'Fanouts' in header: width = 50
            if 'Logic' in header: width = 50
            ws.column_dimensions[chr(64+col) if col <= 26 else 'A'+chr(64+col-26)].width = width

    def save_csv(self, result: Dict[str, Any], base_name: str) -> Path:
        """Save result as flattened CSV (single row)."""
        path = self.output_dir / f"{base_name}.csv"
        
        row = self._flatten_result_to_single_row(result)
        
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=list(row.keys()))
            writer.writeheader()
            writer.writerow(row)
        
        return path

    def _flatten_result_to_single_row(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """Flatten entire result into a single dictionary (one row)."""
        fanout_out = result.get('fanout_output', {})
        validation = result.get('validation', {})
        
        # Format Fanouts List
        valid_fanouts = fanout_out.get('fanouts', [])
        formatted_fanouts = []
        for f in valid_fanouts:
            if isinstance(f, dict):
                formatted_fanouts.append(f"{f.get('query', '')} ({f.get('mapped_type', 'No Type')})")
            else:
                formatted_fanouts.append(str(f))
        
        # Format Rejected Fanouts
        rejected = fanout_out.get('rejected_fanouts', [])
        formatted_rejected = []
        for r in rejected:
            if isinstance(r, dict):
                formatted_rejected.append(f"{r.get('query', '')}: {r.get('reason', '')}")
            else:
                formatted_rejected.append(str(r))
                
        # Format Validation Steps
        val_steps = validation.get('steps', [])
        formatted_validation = []
        for step in val_steps:
            status = "✅" if step.get('is_valid') else "❌"
            formatted_validation.append(f"{status} {step.get('step')}: Score {step.get('score', 0)}")
            if step.get('errors'):
                formatted_validation.append(f"   Errors: {'; '.join(step['errors'])}")
            if step.get('warnings'):
                formatted_validation.append(f"   Warnings: {'; '.join(step['warnings'])}")
        
        return {
            "Main Query": result.get('query', ''),
            "Mode": result.get('mode', ''),
            "Valid Fanouts": "\n".join(formatted_fanouts),
            "Rejected Fanouts": "\n".join(formatted_rejected),
            "Must Cover": "\n".join(fanout_out.get('must_cover', [])),
            "Logic/Reasoning": fanout_out.get('reasoning', '') or fanout_out.get('validation_logic', '') or fanout_out.get('brainstorming_process', ''),
            "Copy Text": result.get('final_copy', ''),
            "Fanout Guardrail": result.get('fanout_guardrail_verdict', ''),
            "Copy Guardrail": result.get('copy_guardrail_verdict', ''),
            "Total Retries": result.get('fanout_retries', 0) + result.get('copy_retries', 0),
            "Quality Score": validation.get('overall_score', 0),
            "Validation Log": "\n".join(formatted_validation),
            "Time (s)": f"{result.get('total_time', 0):.2f}"
        }


===============================================================================================================================================================================


"""
Copy Block Generator CLI

Provides commands for single keyword and batch processing of copy generation.
Outputs both CSV and JSON formats with validation scores.
"""

import argparse
import sys
import logging
import json
import csv
import time
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

from .config import load_config, Strategies
from .utils.logging import setup_logging
from .core.workflow import CopyBlockWorkflow
from .utils.mocks import apply_mocks

logger = logging.getLogger(__name__)


def _load_item_types(config: Dict[str, Any]) -> List[str]:
    """Helper to load item types from config path."""
    item_types = []
    it_path = config.get('data_paths', {}).get('item_types')
    if it_path:
        path_obj = Path(it_path)
        if path_obj.exists():
            try:
                with open(path_obj, 'r') as f:
                    item_types = json.load(f)
                logger.info(f"Loaded {len(item_types)} item types from {it_path}")
            except Exception as e:
                logger.warning(f"Failed to parse item types from {it_path}: {e}")
        else:
            logger.warning(f"Item types file not found at: {it_path}")
    else:
        logger.warning("No item_types path configured in data_paths")
    return item_types


def run_single_keyword(args):
    """Run pipeline for a single keyword."""
    config = load_config(args.config)
    log_file, _ = setup_logging(args.log_dir)
    
    keyword = args.keyword
    mode = args.mode
    
    logger.info(f"Processing keyword: {keyword} [Mode: {mode}]")
    print(f"Processing keyword: {keyword} [Mode: {mode}]")
    
    if args.mock:
        apply_mocks()
        logger.info("Mock mode enabled")

    # Load item types
    item_types = _load_item_types(config)

    # Run Workflow
    workflow = CopyBlockWorkflow(config, mode=mode)
    result = workflow.run(query=keyword, item_types=item_types)
    
    # Add metadata
    result['timestamp'] = datetime.now().isoformat()
    result['log_file'] = str(log_file)
    
    # Output to console
    print(json.dumps(result, indent=2, default=str))
    
    # Save to files
    output_dir = args.output_dir or "output"
    from .utils.output_formatter import OutputFormatter
    formatter = OutputFormatter(output_dir)
    saved_paths = formatter.save_all(result)
    
    print(f"\n{'='*60}")
    print("Output files saved:")
    for fmt, path in saved_paths.items():
        print(f"  {fmt.upper()}: {path}")
    print(f"  LOG: {log_file}")
    print(f"{'='*60}")
    logger.info(f"Saved outputs to {output_dir}")


def run_batch(args):
    """Run pipeline for a batch of keywords with dual CSV/JSON output."""
    config = load_config(args.config)
    log_file, _ = setup_logging(args.log_dir)
    
    if args.mock:
        apply_mocks()
        logger.info("Mock mode enabled")
    
    # Read inputs
    input_file = args.input_file
    if not input_file:
        input_file = config.get('batch_processing', {}).get('input_file')
        if input_file:
            logger.info(f"Using input file from config: {input_file}")
            print(f"Using input file from config: {input_file}")
            
    if not input_file:
        logger.error("No input file provided (args or config)")
        print("Error: No input file provided. Please specify file or set batch_processing.input_file in conf.yml")
        sys.exit(1)
        
    keywords = _read_keywords(str(input_file))
    if not keywords:
        logger.error("No keywords found in input file")
        print("Error: No keywords found in input file")
        sys.exit(1)
    
    # Load item types once
    item_types = _load_item_types(config)
    
    logger.info(f"Starting batch processing for {len(keywords)} keywords [Mode: {args.mode}]")
    print(f"Starting batch processing for {len(keywords)} keywords [Mode: {args.mode}]...")
    
    # Prepare output
    output_dir = Path(args.output_dir or "output")
    output_dir.mkdir(parents=True, exist_ok=True)
    from .utils.output_formatter import OutputFormatter
    formatter = OutputFormatter(str(output_dir))
    
    # Concurrency Config
    batch_cfg = config.get('batch_processing', {})
    parallel_cfg = batch_cfg.get('parallel_processing', {})
    params = {
        'use_parallel': parallel_cfg.get('enabled', False),
        'max_workers': parallel_cfg.get('max_workers', 5)
    }

    full_results = []
    start_time = time.time()
    
    try:
        if params['use_parallel'] and len(keywords) > 1:
            import concurrent.futures
            print(f"🚀 Parallel mode enabled: {params['max_workers']} workers")
            logger.info(f"Parallel mode enabled: {params['max_workers']} workers")
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=params['max_workers']) as executor:
                # Submit all tasks
                future_to_kw = {
                    executor.submit(_process_batch_item, kw, config, args.mode, item_types, formatter): kw 
                    for kw in keywords
                }
                
                # Process as they complete
                completed = 0
                for future in concurrent.futures.as_completed(future_to_kw):
                    kw = future_to_kw[future]
                    completed += 1
                    try:
                        res = future.result()
                        full_results.append(res)
                        
                        # Progress log
                        score = res.get('validation', {}).get('overall_score', 0)
                        status = "✅" if score > 0 else "❌"
                        print(f"[{completed}/{len(keywords)}] {status} {kw} (Score: {score:.1f})")
                        
                    except Exception as e:
                        logger.error(f"Worker failed for {kw}: {e}")
                        
        else:
            # Sequential fallback
            for i, kw in enumerate(keywords, 1):
                print(f"[{i}/{len(keywords)}] Processing: {kw}")
                res = _process_batch_item(kw, config, args.mode, item_types, formatter)
                full_results.append(res)
                
    except KeyboardInterrupt:
        print("\n⚠️ Batch interrupted by user. Saving partial results...")
    
    finally:
        pass
    
    # Save batch report
    print("\nGenerating batch report...")
    batch_paths = formatter.save_batch_report(full_results)
    
    # Summary
    total_time = time.time() - start_time
    success_count = sum(1 for r in full_results if r.get('validation', {}).get('overall_score', 0) > 0)
    
    print(f"\n{'='*60}")
    print(f"Batch complete: {success_count}/{len(keywords)} successful")
    print(f"Total time: {total_time:.1f}s")
    print(f"Batch Reports saved to:")
    for type_, path in batch_paths.items():
        print(f"  {type_.upper()}: {path}")
    print(f"Log file: {log_file}")
    print(f"{'='*60}")
    
    logger.info(f"Batch complete: {success_count}/{len(keywords)} successful in {total_time:.1f}s")


def _process_batch_item(kw: str, config: Dict, mode: str, item_types: List[str], formatter: Any) -> Dict[str, Any]:
    """Process a single keyword item for batch runner."""
    try:
        workflow = CopyBlockWorkflow(config, mode=mode)
        res = workflow.run(query=kw, item_types=item_types)
        
        # Add metadata
        res['timestamp'] = datetime.now().isoformat()
        
        # Save individual backup
        formatter.save_all(res)
        
        return res
        
    except Exception as e:
        logger.error(f"Failed {kw}: {e}", exc_info=True)
        return {
            "query": kw,
            "mode": mode, 
            "validation": {"overall_score": 0},
            "fanout_output": {
                "reasoning": f"FAILED: {str(e)}",
                "rejected_fanouts": [{"query": "ALL", "reason": str(e)}]
            }
        }


def _read_keywords(input_file: str) -> List[str]:
    """Read keywords from CSV or TXT file."""
    keywords = []
    
    if input_file.endswith('.csv'):
        with open(input_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
            # Skip header if present
            if rows and rows[0][0].lower() in ['keyword', 'query', 'main keyword']:
                rows = rows[1:]
            keywords = [r[0].strip() for r in rows if r and r[0].strip()]
    else:
        with open(input_file, 'r', encoding='utf-8') as f:
            keywords = [line.strip() for line in f if line.strip()]
    
    return keywords


def main():
    parser = argparse.ArgumentParser(
        prog="copy_gen",
        description="Copy Block Generator CLI - Generate SEO copy from keywords"
    )
    parser.add_argument("--config", default="conf.yml", help="Path to config file")
    parser.add_argument("--log-dir", default="logs", help="Log directory")
    parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose output")
    
    subparsers = parser.add_subparsers(dest="command", required=True)
    
    # Run command (single keyword)
    run_parser = subparsers.add_parser("run", help="Run generation for a single keyword")
    run_parser.add_argument("keyword", help="Search keyword")
    run_parser.add_argument("--mode", default=Strategies.V1, choices=Strategies.ALL,
                           help="Generation strategy (v1=constrained, v2=unconstrained+filter)")
    run_parser.add_argument("--output-dir", "-o", default="output",
                           help="Output directory for JSON, CSV, and Excel files")
    run_parser.add_argument("--mock", action="store_true", help="Mock API calls for testing")
    run_parser.set_defaults(func=run_single_keyword)
    
    # Batch command
    batch_parser = subparsers.add_parser("batch", help="Run batch generation from file")
    batch_parser.add_argument("input_file", nargs="?", help="Input CSV/TXT file with keywords. Optional if configured in conf.yml")
    batch_parser.add_argument("--output-dir", "-o", default="output",
                             help="Output directory for results")
    batch_parser.add_argument("--mode", default=Strategies.V1, choices=Strategies.ALL)
    batch_parser.add_argument("--mock", action="store_true", help="Mock API calls for testing")
    batch_parser.set_defaults(func=run_batch)
    
    args = parser.parse_args()
    
    # Set verbose logging if requested
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    try:
        args.func(args)
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
        sys.exit(130)
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

